import customtkinter as ctk
import os
import platform
import subprocess
from tkinter import messagebox
from PIL import Image

# Tambahan modul baru untuk membaca berbagai dokumen
import fitz  # PyMuPDF untuk PDF
from docx import Document
from openpyxl import load_workbook

# ======================================================
# UTILITY FUNCTIONS
# ======================================================

def get_file_size_human(size_bytes):
    """Mengubah ukuran byte menjadi format yang mudah dibaca (KB, MB, GB)."""
    if size_bytes < 1024:
        return f"{size_bytes} Bytes"
    elif size_bytes < 1024 * 1024:
        return f"{size_bytes / 1024:.2f} KB"
    elif size_bytes < 1024 * 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f} MB"
    else:
        return f"{size_bytes / (1024 * 1024 * 1024):.2f} GB"


# ======================================================
# MAIN PREVIEW FUNCTION
# ======================================================

class PreviewWindow(ctk.CTkToplevel):
    def __init__(self, parent, file_path):
        super().__init__(parent)
        self.file_path = file_path
        self.parent = parent
        self.title("Preview File Duplikat")
        self.geometry("700x500")
        self.resizable(False, False)

        # Mencegah interaksi dengan window utama
        self.grab_set()

        if not os.path.exists(file_path):
            messagebox.showerror("Error", f"File tidak ditemukan: {file_path}")
            self.destroy()
            return

        self.file_name = os.path.basename(file_path)
        self.file_dir = os.path.dirname(file_path)
        self.file_ext = os.path.splitext(self.file_name)[1].lower()
        self.file_size = os.path.getsize(file_path)

        self.build_ui()

    def build_ui(self):
        # Frame Judul dan Info
        info_frame = ctk.CTkFrame(self, fg_color="transparent")
        info_frame.pack(fill="x", padx=20, pady=(15, 10))

        # Judul (Nama File)
        ctk.CTkLabel(
            info_frame,
            text=self.file_name,
            font=("Poppins", 16, "bold"),
            wraplength=650,
        ).pack(anchor="w")

        # Info Path dan Ukuran
        ctk.CTkLabel(
            info_frame,
            text=f"Path: {self.file_dir}",
            font=("Inter", 10),
            wraplength=650,
        ).pack(anchor="w")

        ctk.CTkLabel(
            info_frame,
            text=f"Ukuran: {get_file_size_human(self.file_size)}",
            font=("Inter", 10),
        ).pack(anchor="w")

        # Area Konten Preview
        content_frame = ctk.CTkFrame(self, fg_color=("white", "#2B2B40"))
        content_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.show_content_preview(content_frame)

    def show_content_preview(self, container):
        """Menampilkan konten file (gambar, teks, PDF, Word, Excel) di container."""

        ext = self.file_ext

        # === Preview Gambar ===
        if ext in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.ico']:
            try:
                img = Image.open(self.file_path)
                max_size = (400, 400)
                img.thumbnail(max_size)
                ctk_img = ctk.CTkImage(light_image=img, dark_image=img, size=img.size)
                ctk.CTkLabel(container, text="", image=ctk_img).pack(expand=True)
            except Exception as e:
                ctk.CTkLabel(container, text=f"Gagal memuat gambar: {e}", text_color="red").pack(pady=50)

        # === Preview File Teks ===
        elif ext in ['.txt', '.log', '.csv', '.py', '.html', '.css', '.js']:
            try:
                with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    preview_text = f.read(2000)
                textbox = ctk.CTkTextbox(container, wrap="word", width=600, height=400)
                textbox.insert("1.0", preview_text + "\n\n--- [Hanya sebagian isi file] ---")
                textbox.configure(state="disabled")
                textbox.pack(padx=10, pady=10, fill="both", expand=True)
            except Exception as e:
                ctk.CTkLabel(container, text=f"Gagal membaca teks: {e}", text_color="red").pack(pady=50)

        # === Preview PDF ===
        elif ext == ".pdf":
            try:
                doc = fitz.open(self.file_path)
                text = ""
                for page in doc:
                    text += page.get_text()
                    if len(text) > 2000:
                        break
                textbox = ctk.CTkTextbox(container, wrap="word", width=600, height=400)
                textbox.insert("1.0", text.strip() + "\n\n--- [Potongan isi PDF] ---")
                textbox.configure(state="disabled")
                textbox.pack(padx=10, pady=10, fill="both", expand=True)
            except Exception as e:
                ctk.CTkLabel(container, text=f"Gagal membaca PDF: {e}", text_color="red").pack(pady=50)

        # === Preview Word ===
        elif ext == ".docx":
            try:
                doc = Document(self.file_path)
                text = "\n".join([p.text for p in doc.paragraphs[:50]])
                textbox = ctk.CTkTextbox(container, wrap="word", width=600, height=400)
                textbox.insert("1.0", text.strip() + "\n\n--- [Potongan isi dokumen Word] ---")
                textbox.configure(state="disabled")
                textbox.pack(padx=10, pady=10, fill="both", expand=True)
            except Exception as e:
                ctk.CTkLabel(container, text=f"Gagal membaca Word: {e}", text_color="red").pack(pady=50)

        # === Preview Excel ===
        elif ext in [".xls", ".xlsx"]:
            try:
                wb = load_workbook(self.file_path, read_only=True)
                sheet = wb.active
                text = ""
                for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
                    text += "\t".join(str(cell) if cell is not None else "" for cell in row) + "\n"
                textbox = ctk.CTkTextbox(container, wrap="none", width=600, height=400)
                textbox.insert("1.0", text.strip() + "\n\n--- [Potongan isi Excel] ---")
                textbox.configure(state="disabled")
                textbox.pack(padx=10, pady=10, fill="both", expand=True)
            except Exception as e:
                ctk.CTkLabel(container, text=f"Gagal membaca Excel: {e}", text_color="red").pack(pady=50)

        # === Format lain (tidak didukung) ===
        else:
            ctk.CTkLabel(
                container,
                text=f"Tidak ada preview untuk format '{self.file_ext}'.",
                font=("Poppins", 14),
                text_color="#B3B3B3"
            ).pack(pady=50)


# ======================================================
# DIPANGGIL DARI main_window.py
# ======================================================
def show_preview_window(parent, file_path):
    if not os.path.exists(file_path):
        messagebox.showerror("Error", f"File tidak ditemukan: {file_path}")
        return
    PreviewWindow(parent, file_path)
